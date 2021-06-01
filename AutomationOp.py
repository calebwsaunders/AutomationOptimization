#! AutomationOp.py - A simple python script to take the data from a dispense
#  report from MHSG and make it more available for aligning a site's automation
#  with their dispensing history.
#  python AutomationOp.py

import openpyxl
import glob
import calendar
import pyinputplus as pyip
from openpyxl.styles import Font
from datetime import date

# A variable for assigning the year to the output file
data_year = date.today().year
data_month = date.today().month
month_abbr = calendar.month_abbr[data_month]

# Using glob to make a list of all the excel files in the current directory.
files = glob.glob('*.xlsx')

# Making an array to accept an array for each line, so each occurrence, from each excel document.
occurrences = []

# Making two arrays to differentiate between new Rxs and refills for larger sites with separated refills.
new_occurrences = []
refill_occurrences = []

# Asking the customer is they want to separate new from refills.
print('\nDo you want to combine new prescriptions and refills ("Yes" to combine, "No" to separate?')
new_refill_request = pyip.inputMenu(['Yes', 'No'], numbered=True)

# Ask the customer what they want.
# The numbered output will be:
# 1 = all medications
# 2 = only legend drugs
# 3 = legend drugs and CIII-Vs
# 4 = all controlled medications
# 5 = minor controls
# 6 = CIIs
print('\nWhich class of medications are you optimizing?')
data_request = pyip.inputMenu(['Both controlled and non-controlled medications', 'Only non-controlled medications',
                               'Both non-controlled and CIII-Vs', 'Only all controlled medications', 'Only CIII-Vs',
                               'Only CIIs'], numbered=True)

# Making arrays for choosing between controlled medications or non-controlled medications.
all_controlled_medications = [2, '2N', 3, '3N', 4, 5]
minor_controlled_medications = [3, '3N', 4, 5]
narcotics = [2, '2N']

# Making a list of items to exclude if found in the item description.
excluding_criteria = ['mL', 'g']
excluded_items = ['inhaler', 'patch', 'packet', 'cream', 'gel', 'solution', 'suppository', '(21)',
                  '(28)', '(91)', '(60EA)', '(53EA)', '(56EA)', '(16EA)', '(6EA)', '(3EA)', '[100EA]',
                  '[60EA]', '[50EA]', '[30EA]', '[25EA]', '[21EA]', '[16EA]', '[8EA]', '[6EA]', '[4EA]',
                  '[3EA]', '[2EA]', '[1EA]', '[3PENS]', '[2PENS]', 'ring', 'needle', 'lancet', 'Monitor',
                  'spray', '(60EA/30Dose)', '(60EA/30Dos)', 'test strip', 'Inj', 'Kit', 'Ring',
                  'anastrozole', 'letrozole', 'methotrexate', 'tamoxifen']


def get_item_id(occurrence):
    """Going to return the item id from each occurrence to be used to sort the occurrences."""
    item_id = occurrence[0]
    return item_id

def check_for_excludable_items(occurrence):
    """Searching the description of each occurrence and checking against all excludable items.
    If the descriptions contains an excludable item the output will be set to True."""
    output = False
    for item in excluded_items:
        if (occurrence[1].__contains__(item)):
            output = True
            break
        else:
            continue
    return output

def append_to_occurrences(row):
    global occurrences
    item_id = sheet[f'B{row}'].value
    description = sheet[f'C{row}'].value
    dispenses = sheet[f'G{row}'].value
    quantity = sheet[f'H{row}'].value
    occurrence = [item_id, description, dispenses, quantity]
    occurrences.append(occurrence)
    return

def append_to_new_or_refill_occurrences(row):
    global new_occurrences
    global refill_occurrences
    if sheet[f'F{row}'].value == 'New Rx':
        item_id = sheet[f'B{row}'].value
        description = sheet[f'C{row}'].value
        dispenses = sheet[f'G{row}'].value
        quantity = sheet[f'H{row}'].value
        occurrence = [item_id, description, dispenses, quantity]
        new_occurrences.append(occurrence)
    else:
        item_id = sheet[f'B{row}'].value
        description = sheet[f'C{row}'].value
        dispenses = sheet[f'G{row}'].value
        quantity = sheet[f'H{row}'].value
        occurrence = [item_id, description, dispenses, quantity]
        refill_occurrences.append(occurrence)
    return

# Setting up a loop to iterate over every excel file identified above.
for file in files:
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    # Setting up a loop to iterate over every row in the excel document; will grab the item id (B), description (C),
    # dispenses (G), and quantity(H). Starting at 2 to skip the header information and adding + 1 to max_row
    # in order to get the last item of each sheet.
    for row in range(2, sheet.max_row + 1):
        if sheet[f'I{row}'].value in excluding_criteria:
            continue
        else:
            # User selected #1
            if data_request == 'Both controlled and non-controlled medications':
                if new_refill_request == 'Yes':
                    append_to_occurrences(row)
                else:
                    append_to_new_or_refill_occurrences(row)

            # User selected #2
            elif data_request == 'Only non-controlled medications':
                if sheet[f'D{row}'].value in all_controlled_medications:
                    continue
                else:
                    if new_refill_request == 'Yes':
                        append_to_occurrences(row)
                    else:
                        append_to_new_or_refill_occurrences(row)

            # User selected #3
            elif data_request == 'Both non-controlled and CIII-Vs':
                if sheet[f'D{row}'].value in narcotics:
                    continue
                else:
                    if new_refill_request == 'Yes':
                        append_to_occurrences(row)
                    else:
                        append_to_new_or_refill_occurrences(row)

            # User selected #4
            elif data_request == 'Only all controlled medications':
                if sheet[f'D{row}'].value in all_controlled_medications:
                    if new_refill_request == 'Yes':
                        append_to_occurrences(row)
                    else:
                        append_to_new_or_refill_occurrences(row)
                else:
                    continue

            # User selected #5
            elif data_request == 'Only CIII-Vs':
                if sheet[f'D{row}'].value in minor_controlled_medications:
                    if new_refill_request == 'Yes':
                        append_to_occurrences(row)
                    else:
                        append_to_new_or_refill_occurrences(row)
                else:
                    continue

            # User selected #6
            elif data_request == 'Only CIIs':
                if sheet[f'D{row}'].value in narcotics:
                    if new_refill_request == 'Yes':
                        append_to_occurrences(row)
                    else:
                        append_to_new_or_refill_occurrences(row)
                else:
                    continue

    # Giving some feedback while running program and closing the workbook before starting the next one.
    print(f'Finished processing {file}')
    wb.close()

    # Excluding medications and supplies that cannot be placed in automation, then adding pertinent
    # items to a new list.
    pertinent_occurrences = []         # For combined list.
    pertinent_new_occurrences = []     # For new occurrences when separating.
    pertinent_refill_occurrences = []  # For refill occurrences when separating.

    if new_refill_request == 'Yes':
        for occurrence in occurrences:
            if check_for_excludable_items(occurrence):
                continue
            else:
                pertinent_occurrences.append(occurrence)

        # Sorting the data before we can count the times dispensed and quantity dispensed for each item id.
        sorted_occurrences = sorted(pertinent_occurrences, key=get_item_id)

        # Setting up variables to start identifying multiple occurrences and adding into one output value.
        item_id = sorted_occurrences[0][0]
        description = sorted_occurrences[0][1]
        dispenses = 0
        quantity = 0

        # Open new workbook to write to and set active sheet.
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Setting up font variables for the title and headers
        fontObjTitle = Font(name='Times New Roman', size=16, bold=True)
        fontObjHeader = Font(name='Times New Roman', size=14)

        # Formatting and writing to the header.
        sheet.oddHeader.center.text = f'Automation Optimization: {month_abbr}{data_year}'
        sheet.oddHeader.center.size = 16
        sheet.oddHeader.center.font = 'Times New Roman'

        # Formatting the columns
        sheet['A1'].font = fontObjHeader
        sheet['B1'].font = fontObjHeader
        sheet['C1'].font = fontObjHeader
        sheet['D1'].font = fontObjHeader

        # Writing the header information
        sheet['A1'] = 'Item ID'
        sheet['B1'] = 'Description'
        sheet['C1'] = 'Dispenses'
        sheet['D1'] = 'Quantity Dispensed'

        row_to_write_to = 2

        # Loop through all occurrences in sorted occurrence. While item_id is the same add to the current dispenses
        # and quantity variables.
        for occurrence in sorted_occurrences:
            if occurrence[0] == item_id:
                dispenses += occurrence[2]
                quantity += occurrence[3]
            else:
                # Write values to new workbook
                sheet[f'A{row_to_write_to}'] = item_id
                sheet[f'B{row_to_write_to}'] = description
                sheet[f'C{row_to_write_to}'] = dispenses
                sheet[f'D{row_to_write_to}'] = quantity

                # Increment the row to write to next
                row_to_write_to += 1

                # Set new variables with the next values
                if occurrence[0] is not None:
                    item_id = occurrence[0]
                    description = occurrence[1]
                    dispenses = occurrence[2]
                    quantity = occurrence[3]
                else:
                    break

        wb.save(f'Automation Optimization Review {month_abbr}{data_year}.xlsx')
        wb.close()

    else:
        for occurrence in new_occurrences:
            if check_for_excludable_items(occurrence):
                continue
            else:
                pertinent_new_occurrences.append(occurrence)

        for occurrence in refill_occurrences:
            if check_for_excludable_items(occurrence):
                continue
            else:
                pertinent_refill_occurrences.append(occurrence)

        # Sorting the data before we can count the times dispensed and quantity dispensed for each item id.
        sorted_new_occurrences = sorted(pertinent_new_occurrences, key=get_item_id)
        sorted_refill_occurrences = sorted(pertinent_refill_occurrences, key=get_item_id)

        # Setting up variables to start identifying multiple occurrences and adding into one output value.
        item_id_new = sorted_new_occurrences[0][0]
        description_new = sorted_new_occurrences[0][1]
        dispenses_new = 0
        quantity_new = 0

        item_id_refill = sorted_refill_occurrences[0][0]
        description_refill = sorted_refill_occurrences[0][1]
        dispenses_refill = 0
        quantity_refill = 0

        # Open new workbook to write to and set active sheet.
        wb = openpyxl.Workbook()
        sheet_new = wb.create_sheet('New', 0)
        sheet_refills = wb.create_sheet('Refills', 1)

        # Setting up font variables for the title and headers
        fontObjTitle = Font(name='Times New Roman', size=16, bold=True)
        fontObjHeader = Font(name='Times New Roman', size=14)

        # Formatting and writing to the header.
        sheet_new.oddHeader.center.text = f'Automation Optimization: {month_abbr}{data_year}'
        sheet_new.oddHeader.center.size = 16
        sheet_new.oddHeader.center.font = 'Times New Roman'

        sheet_refills.oddHeader.center.text = f'Automation Optimization: {month_abbr}{data_year}'
        sheet_refills.oddHeader.center.size = 16
        sheet_refills.oddHeader.center.font = 'Times New Roman'

        # Formatting the columns
        sheet_new['A1'].font = fontObjHeader
        sheet_new['B1'].font = fontObjHeader
        sheet_new['C1'].font = fontObjHeader
        sheet_new['D1'].font = fontObjHeader

        sheet_refills['A1'].font = fontObjHeader
        sheet_refills['B1'].font = fontObjHeader
        sheet_refills['C1'].font = fontObjHeader
        sheet_refills['D1'].font = fontObjHeader

        # Writing the header information
        sheet_new['A1'] = 'Item ID'
        sheet_new['B1'] = 'Description'
        sheet_new['C1'] = 'Dispenses'
        sheet_new['D1'] = 'Quantity Dispensed'

        sheet_refills['A1'] = 'Item ID'
        sheet_refills['B1'] = 'Description'
        sheet_refills['C1'] = 'Dispenses'
        sheet_refills['D1'] = 'Quantity Dispensed'

        row_to_write_to_new = 2
        row_to_write_to_refills = 2

        # Loop through all occurrences in sorted occurrence. While item_id is the same add to the current dispenses
        # and quantity variables.
        for occurrence in sorted_new_occurrences:
            if occurrence[0] == item_id_new:
                dispenses_new += occurrence[2]
                quantity_new += occurrence[3]
            else:
                # Write values to new workbook
                sheet_new[f'A{row_to_write_to_new}'] = item_id_new
                sheet_new[f'B{row_to_write_to_new}'] = description_new
                sheet_new[f'C{row_to_write_to_new}'] = dispenses_new
                sheet_new[f'D{row_to_write_to_new}'] = quantity_new

                # Increment the row to write to next
                row_to_write_to_new += 1

                # Set new variables with the next values
                if occurrence[0] is not None:
                    item_id_new = occurrence[0]
                    description_new = occurrence[1]
                    dispenses_new = occurrence[2]
                    quantity_new = occurrence[3]
                else:
                    break

        for occurrence in sorted_refill_occurrences:
            if occurrence[0] == item_id_refill:
                dispenses_refill += occurrence[2]
                quantity_refill += occurrence[3]
            else:
                # Write values to new workbook
                sheet_refills[f'A{row_to_write_to_refills}'] = item_id_refill
                sheet_refills[f'B{row_to_write_to_refills}'] = description_refill
                sheet_refills[f'C{row_to_write_to_refills}'] = dispenses_refill
                sheet_refills[f'D{row_to_write_to_refills}'] = quantity_refill

                # Increment the row to write to next
                row_to_write_to_refills += 1

                # Set new variables with the next values
                if occurrence[0] is not None:
                    item_id_refill = occurrence[0]
                    description_refill = occurrence[1]
                    dispenses_refill = occurrence[2]
                    quantity_refill = occurrence[3]
                else:
                    break

        # Cleaning up the sheet before saving.
        sheet_to_delete = wb['Sheet']
        wb.remove_sheet(sheet_to_delete)

        # Save and close the finished product.
        wb.save(f'Automation Optimization Review {month_abbr}{data_year}.xlsx')
        wb.close()